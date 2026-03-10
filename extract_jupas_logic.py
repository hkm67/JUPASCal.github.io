import openpyxl
import json
import os

def extract_jupas_logic(file_path, output_json):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Reference Score Calculation']
    
    # Headers from Col 27 to 48 (1-indexed)
    # Col 2: Code, Col 3: Inst, Col 4: Faculty, Col 5: Name, Col 6: Method
    
    data = []
    
    # Find the last row with data in Col 2 (Code)
    max_row = ws.max_row
    for row in range(13, max_row + 1):
        code = ws.cell(row, 2).value
        if not code or code == "":
            continue
            
        # Basic Info
        entry = {
            "code": str(code).strip(),
            "institution": str(ws.cell(row, 3).value).strip(),
            "faculty": str(ws.cell(row, 4).value).strip(),
            "programme": str(ws.cell(row, 5).value).strip(),
            "formula_2024": str(ws.cell(row, 6).value).strip(),
            "formula_2025": str(ws.cell(row, 29).value).strip(),
            "admission_requirements": [
                str(ws.cell(row, 27).value).strip(),
                str(ws.cell(row, 28).value).strip()
            ],
            "updates": [
                str(ws.cell(row, 30).value).strip(),
                str(ws.cell(row, 31).value).strip()
            ],
            "calc_details": [
                str(ws.cell(row, 32).value).strip(),
                str(ws.cell(row, 33).value).strip()
            ],
            "weightings": [
                str(ws.cell(row, 34).value).strip(),
                str(ws.cell(row, 35).value).strip(),
                str(ws.cell(row, 36).value).strip()
            ],
            "other_considerations": [
                str(ws.cell(row, 37).value).strip(),
                str(ws.cell(row, 38).value).strip(),
                str(ws.cell(row, 39).value).strip()
            ],
            "extra_data": [
                str(ws.cell(row, 40).value).strip(),
                str(ws.cell(row, 41).value).strip(),
                str(ws.cell(row, 42).value).strip(),
                str(ws.cell(row, 43).value).strip(),
                str(ws.cell(row, 44).value).strip(),
                str(ws.cell(row, 45).value).strip()
            ],
            "estimates": [
                str(ws.cell(row, 46).value).strip(),
                str(ws.cell(row, 47).value).strip()
            ],
            "retake_penalty": str(ws.cell(row, 48).value).strip()
        }
        
        # Cleanup "/" and "None" to null for cleaner JSON
        def clean_val(v):
            if v in ["/", "None", "None None", "None/None", "None / None", "None\nNone"]:
                return None
            return v
            
        for key in entry:
            if isinstance(entry[key], list):
                entry[key] = [clean_val(i) for i in entry[key]]
                # Remove trailing Nones
                while entry[key] and entry[key][-1] is None:
                    entry[key].pop()
            else:
                entry[key] = clean_val(entry[key])
                
        data.append(entry)
        
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"Successfully extracted {len(data)} programmes to {output_json}")

if __name__ == "__main__":
    xl_file = "Archives/2025 JUPAS 計分器/下載嚟用 (v1.0.3) 2025 JUPAS Cal.xlsx"
    extract_jupas_logic(xl_file, "jupas_2025_logic.json")
